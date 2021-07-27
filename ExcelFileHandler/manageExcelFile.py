import xlsxwriter
import openpyxl
from openpyxl.styles import Alignment


class ManageExcelFile:

    def save_into_xls(self, dic, i):
        wb = openpyxl.load_workbook(filename=self.filename)
        ws = wb.get_sheet_by_name('Sheet1')

        sheet = wb.worksheets[0]
        row = sheet.max_row + 1
        col = 2

        ws.cell(row=row, column=1, value=i)
        for key, value in dic.items():
            if key == 'overview':
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=col, value=value)

            col = col + 1

        wb.save(self.filename)
        wb.close()

    def create_xls_file(self):
        workbook = xlsxwriter.Workbook(self.filename)
        worksheet = workbook.add_worksheet()

        bold = workbook.add_format({'bold': True})

        worksheet.write(1, 0, '#', bold)
        worksheet.write(1, 1, 'Name', bold)
        worksheet.write(1, 2, 'Overview', bold)
        worksheet.write(1, 3, 'Website', bold)
        worksheet.write(1, 4, 'Size', bold)
        worksheet.write(1, 5, 'Industry', bold)
        worksheet.write(1, 6, 'Headquarters', bold)
        worksheet.write(1, 7, 'Type', bold)
        worksheet.write(1, 8, 'Founded', bold)
        worksheet.write(1, 9, 'Specialties', bold)

        workbook.close()

    def __init__(self, filename):
        self.filename = filename
        self.create_xls_file()
